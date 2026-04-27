"""
Batch-updates content/projects/*.json:
  - Sets name (all variants) to the Initiative Name from the mapping CSV
  - Replaces old name occurrences in text fields with the new name
  - Adds entityDescription, entityRole, bizModel, entitySize, geoReach,
    sectorFocus, resourceType, caseCode fields from the coding cases Excel
  - Moves non-entity files to content/non-projects/
  - Writes edge_cases.xlsx with two sheets of exception lists
"""

import json
import csv
import re
import shutil
import openpyxl
from pathlib import Path

# ── Label look-ups (full labels from the variable definitions) ──────────────

V1_LABELS = {
    'FP': 'Primary Production',
    'FT': 'Transformation/Processing',
    'FD': 'Distribution/Logistics',
    'FR': 'Retail',
    'FS': 'Food Service/Hospitality',
    'FW': 'Waste & Circular Economy',
    'SS': 'Support Services',
    'FI': 'Finance & Investment',
    'RG': 'Governance & Regulation',
    'CN': 'Civil Society & Networks',
    'RO': 'Other Role',
}

V2_LABELS = {
    'PR': 'For-profit private',
    'CO': 'Cooperative/mutual',
    'SO': 'Social enterprise',
    'NP': 'Non-profit/charity',
    'PU': 'Public body',
    'PP': 'Public-private partnership',
    'IF': 'Individual/informal',
    'BX': 'Other model',
}

V3_LABELS = {
    'IN': 'Individual/Solo',
    'MI': 'Micro',
    'SM': 'Small',
    'ME': 'Medium',
    'LA': 'Large',
    'VL': 'Very Large',
    'UK': 'Unknown',
}

V4_LABELS = {
    'LO': 'Local',
    'RE': 'Regional',
    'NAT': 'National',
    'INT': 'International',
    'GX': 'Unknown',
}

V5_LABELS = {
    'AGR': 'Agriculture/Agroecology',
    'FSH': 'Fisheries & Aquaculture',
    'LST': 'Livestock & Dairy',
    'FOD': 'Food Processing/Products',
    'DIS': 'Distribution & Logistics',
    'WST': 'Waste & Circular Economy',
    'ENE': 'Energy',
    'WAT': 'Water',
    'TEC': 'Technology & Innovation',
    'HLT': 'Health & Nutrition',
    'ENV': 'Environment & Biodiversity',
    'PLY': 'Policy & Advocacy',
    'SOC': 'Social & Community',
    'SX': 'Other sector',
}

V6_LABELS = {
    'GD': 'Goods & Products',
    'SV': 'Services',
    'FN': 'Financial',
    'AC': 'Access & Infrastructure',
    'KN': 'Knowledge & Information',
    'LG': 'Licences & Permissions',
    'RS': 'Restriction & Obligation',
    'SP': 'Support & Endorsement',
    'RX': 'Other resource',
}


def codes_to_labels(codes_str, label_map):
    """Parse comma-separated codes → list of full labels."""
    if not codes_str:
        return []
    return [label_map[c.strip()] for c in codes_str.split(',')
            if c.strip() in label_map]


def replace_in_value(value, old, new):
    """Recursively replace old → new in any JSON value (str/dict/list)."""
    if not old or old == new:
        return value
    if isinstance(value, str):
        return re.sub(re.escape(old), new, value, flags=re.IGNORECASE)
    if isinstance(value, dict):
        return {k: replace_in_value(v, old, new) for k, v in value.items()}
    if isinstance(value, list):
        return [replace_in_value(item, old, new) for item in value]
    return value


# ── Read source files ────────────────────────────────────────────────────────

def read_mapping(csv_path):
    """Returns (pubid_map, code_to_pubid):
      pubid_map    : {pubId(int) -> {'code': str, 'name': str}}
      code_to_pubid: {initiative_code(str) -> pubId(int) | None}
    """
    pubid_map = {}
    code_to_pubid = {}
    with open(csv_path, encoding='utf-8-sig', newline='') as f:
        for row in csv.DictReader(f):
            code = row['Initiative Code'].strip()
            name = row['Initiative Name'].strip()
            raw_id = row.get('website id', '').strip()
            pub_id = None
            if raw_id:
                try:
                    pub_id = int(float(raw_id))
                except ValueError:
                    pass
            if pub_id is not None:
                pubid_map[pub_id] = {'code': code, 'name': name}
            code_to_pubid[code] = pub_id
    return pubid_map, code_to_pubid


def read_coding_cases(xlsx_path):
    """Returns {case_code(str) -> dict with description + V1-V6 label lists/strings}."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb['Case Entity Profiles']
    rows = list(ws.iter_rows(values_only=True))
    coding = {}
    for row in rows[1:]:          # skip header
        if not row[0]:
            continue
        code = str(row[0]).strip()
        v1_raw = str(row[4]).strip() if row[4] else ''
        v2_raw = str(row[6]).strip() if row[6] else ''
        v3_raw = str(row[8]).strip() if row[8] else ''
        v4_raw = str(row[10]).strip() if row[10] else ''
        v5_raw = str(row[12]).strip() if row[12] else ''
        v6_raw = str(row[14]).strip() if row[14] else ''

        v2_labels = codes_to_labels(v2_raw, V2_LABELS)
        v3_labels = codes_to_labels(v3_raw, V3_LABELS)
        v4_labels = codes_to_labels(v4_raw, V4_LABELS)
        v6_labels = codes_to_labels(v6_raw, V6_LABELS)

        coding[code] = {
            'description': str(row[3]).strip() if row[3] else '',
            'entityRole':    codes_to_labels(v1_raw, V1_LABELS),
            'bizModel':      v2_labels[0] if v2_labels else '',
            'entitySize':    v3_labels[0] if v3_labels else '',
            'geoReach':      v4_labels[0] if v4_labels else '',
            'sectorFocus':   codes_to_labels(v5_raw, V5_LABELS),
            'resourceType':  v6_labels[0] if v6_labels else '',
        }
    return coding


# ── Main processing ──────────────────────────────────────────────────────────

def process(base_dir: Path):
    mapping, code_to_pubid = read_mapping(base_dir / 'Global tracker - Mapping.csv')
    coding = read_coding_cases(base_dir / 'coding cases.xlsx')

    projects_dir   = base_dir / 'content' / 'projects'
    non_proj_dir   = base_dir / 'content' / 'non-projects'
    non_proj_dir.mkdir(parents=True, exist_ok=True)

    moved, updated, errors = [], [], []

    for json_file in sorted(projects_dir.glob('*.json')):
        try:
            with open(json_file, encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            errors.append((json_file.name, str(e)))
            continue

        pub_id = data.get('pubId')

        if pub_id not in mapping:
            shutil.move(str(json_file), str(non_proj_dir / json_file.name))
            moved.append(json_file.name)
            continue

        # ── Entity: update in-place ───────────────────────────────────────
        entry     = mapping[pub_id]
        init_code = entry['code']
        new_name  = entry['name']

        # old name (used for text replacement)
        if isinstance(data.get('name'), str):
            old_name = data['name']
        elif isinstance(data.get('name'), dict):
            old_name = data['name'].get('en', '')
        else:
            old_name = ''

        # Replace all language variants with the canonical mapping name
        data['name'] = new_name

        # Fix name occurrences in text fields
        if old_name and old_name.strip().lower() != new_name.strip().lower():
            for field in ('description', 'about', 'timeline', 'offers', 'lookingFor', 'gallery'):
                if field in data:
                    data[field] = replace_in_value(data[field], old_name, new_name)

        # Entity coding fields
        data['caseCode'] = init_code
        if init_code in coding:
            c = coding[init_code]
            data['entityDescription'] = c['description']
            data['entityRole']   = c['entityRole']
            data['bizModel']     = c['bizModel']
            data['entitySize']   = c['entitySize']
            data['geoReach']     = c['geoReach']
            data['sectorFocus']  = c['sectorFocus']
            data['resourceType'] = c['resourceType']

        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        updated.append((json_file.name, init_code, new_name))

    return mapping, coding, code_to_pubid, moved, updated, errors


def write_edge_cases(mapping, coding, code_to_pubid, out_path: Path):
    wb = openpyxl.Workbook()

    # Sheet 1 — Coded but no website id in mapping
    ws1 = wb.active
    ws1.title = 'Coded - No Website ID'
    ws1.append(['Case Code', 'Short Description'])
    for code in sorted(coding):
        if code_to_pubid.get(code) is None:
            desc = coding[code]['description']
            ws1.append([code, desc[:120] if desc else ''])

    # Sheet 2 — In mapping (has website id) but not in coding cases
    ws2 = wb.create_sheet('In Mapping - Not Coded')
    ws2.append(['Initiative Code', 'Initiative Name', 'Website ID'])
    for pub_id, entry in sorted(mapping.items(), key=lambda x: x[1]['code']):
        if entry['code'] not in coding:
            ws2.append([entry['code'], entry['name'], pub_id])

    wb.save(out_path)


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == '__main__':
    base = Path(__file__).parent.parent  # socioscope-website/

    mapping, coding, code_to_pubid, moved, updated, errors = process(base)

    edge_path = base / 'edge_cases.xlsx'
    write_edge_cases(mapping, coding, code_to_pubid, edge_path)

    print(f'Updated : {len(updated)} entity files')
    print(f'Moved   : {len(moved)} files → content/non-projects/')
    print(f'Errors  : {len(errors)}')
    if errors:
        for name, err in errors:
            print(f'  {name}: {err}')
    print(f'Edge cases written to edge_cases.xlsx')
